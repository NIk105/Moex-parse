import smtplib
import time
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formatdate
from email import encoders
from functools import reduce

import numpy
import openpyxl as openpyxl
import pandas
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.select import Select

file_name = "C:/temp/Отчет.xlsx"
sheet_name = "Лист1"


def send_mail(subject, text, mail_to, file, isTls=True, settings=None):
    msg = MIMEMultipart()
    msg['From'] = 'admin.vs@mail.ru'
    msg['To'] = mail_to
    msg['Date'] = formatdate(localtime=True)
    msg['Subject'] = subject
    msg.attach(MIMEText(text))


    part = MIMEBase('application', "octet-stream")
    with open(file, "rb") as f:
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', 'attachment', filename='Отчет.xlsx')
    msg.attach(part)

    with smtplib.SMTP_SSL('smtp.mail.ru', 465) as server:
        server.login('admin.vs@mail.ru', '******')
        server.sendmail('admin.vs@mail.ru', mail_to, msg.as_string())
        print("Successfully sent email")


def isElementPresent(CLASS_NAME, locator):
    try:
        driver.find_element(CLASS_NAME, locator)
    except NoSuchElementException:
        print('No Such Element ', locator)
        return False
    return True


driver = webdriver.Chrome(r'C:\temp\chromedriver.exe')
driver.maximize_window()
driver.get("https://www.moex.com/")
time.sleep(1)

try:

    # find elements
    if isElementPresent("xpath", "//html/body/div[3]/div[2]/div/div/div[2]/nav/span[1]/button"):
        elem = driver.find_element("xpath", "//html/body/div[3]/div[2]/div/div/div[2]/nav/span[1]/button")
        elem.click()
    time.sleep(1)

    if isElementPresent("xpath", "//div[@class='item']//child::a[contains(text(), 'Срочный рынок')]"):
        elem = driver.find_element("xpath", "//div[@class='item']//child::a[contains(text(), 'Срочный рынок')]")
        elem.click()
    time.sleep(1)

    try:
        elem = driver.find_element(by=By.LINK_TEXT, value="Согласен")
        elem.click()
    except NoSuchElementException:
        pass

    # if isElementPresent("xpath", '//*[@id="content_disclaimer"]/div/div/div/div[1]/div/a[1]'):
    #    elem = driver.find_element("xpath", '//*[@id="content_disclaimer"]/div/div/div/div[1]/div/a[1]')
    #    elem.click()
    time.sleep(1)

    if isElementPresent("xpath", '//*[@id="ctl00_frmLeftMenuWrap"]/div/div/div/div[2]/div[13]/a'):
        elem = driver.find_element("xpath", '//*[@id="ctl00_frmLeftMenuWrap"]/div/div/div/div[2]/div[13]/a')
        elem.click()
    time.sleep(1)

    clcik = driver.find_element(By.XPATH, '//*[@id="d1day"]')
    ActionChains(driver).click(clcik).perform()
    ActionChains(driver).send_keys(Keys.PAGE_UP).perform()
    ActionChains(driver).send_keys(Keys.PAGE_UP).perform()
    ActionChains(driver).send_keys(Keys.ENTER).perform()
    clcik = driver.find_element(By.XPATH, '//*[@id="d2day"]')
    ActionChains(driver).click(clcik).perform()
    ActionChains(driver).send_keys(Keys.PAGE_DOWN).perform()
    ActionChains(driver).send_keys(Keys.PAGE_DOWN).perform()
    ActionChains(driver).send_keys(Keys.ENTER).perform()
    clcik = driver.find_element(By.XPATH, '//*[@id="d2month"]')
    ActionChains(driver).click(clcik).perform()
    ActionChains(driver).send_keys(Keys.UP).perform()
    ActionChains(driver).send_keys(Keys.ENTER).perform()

    elem = driver.find_element("xpath", '//*[@id="currency-rate-container"]/form/div[4]/div[2]/div/div[5]/input')
    elem.click()

    currency1 = 'USD/RUB - Доллар США к российскому рублю'
    currency1_short = 'USD/RUB'
    currency2 = 'JPY/RUB - Японская йена к российскому рублю'
    currency2_short = 'JPY/RUB'

    df_dict = {}

    elem = driver.find_element_by_id('ctl00_PageContent_CurrencySelect')
    elem.send_keys(currency1)

    elem = driver.find_element_by_class_name('tablels')
    num_row1 = len(elem.find_elements_by_xpath('//tr[@*]'))
    values1 = [value.text.replace(',', '.') for value in elem.find_elements_by_xpath('//tr[@*]//child::td')]
    num_col1 = int(len(values1) / num_row1)
    data_list1 = numpy.reshape(values1, (num_row1, num_col1))

    df1 = pandas.DataFrame(data_list1)
    df1 = df1[[0, 1, 2]]
    df1.columns = [f'Дата {currency1_short}', f'Курс {currency1_short}', f'Время {currency1_short}']
    df1 = df1.astype({f'Курс {currency1_short}': 'float64'})
    df_dict[currency1] = df1[[f'Дата {currency1_short}', f'Курс {currency1_short}', f'Время {currency1_short}']]

    elem = driver.find_element_by_id('ctl00_PageContent_CurrencySelect')
    elem.send_keys(currency2)

    clcik = driver.find_element(By.XPATH, '//*[@id="d1day"]')
    ActionChains(driver).click(clcik).perform()
    ActionChains(driver).send_keys(Keys.PAGE_UP).perform()
    ActionChains(driver).send_keys(Keys.PAGE_UP).perform()
    ActionChains(driver).send_keys(Keys.ENTER).perform()
    clcik = driver.find_element(By.XPATH, '//*[@id="d2day"]')
    ActionChains(driver).click(clcik).perform()
    ActionChains(driver).send_keys(Keys.PAGE_DOWN).perform()
    ActionChains(driver).send_keys(Keys.PAGE_DOWN).perform()
    ActionChains(driver).send_keys(Keys.ENTER).perform()
    clcik = driver.find_element(By.XPATH, '//*[@id="d2month"]')
    ActionChains(driver).click(clcik).perform()
    ActionChains(driver).send_keys(Keys.UP).perform()
    ActionChains(driver).send_keys(Keys.ENTER).perform()

    elem = driver.find_element_by_class_name('tablels')
    num_row2 = len(elem.find_elements_by_xpath('//tr[@*]'))
    values2 = [value.text.replace(',', '.') for value in elem.find_elements_by_xpath('//tr[@*]//child::td')]
    num_col2 = int(len(values2) / num_row2)
    data_list2 = numpy.reshape(values2, (num_row2, num_col2))
    df2 = pandas.DataFrame(data_list2)
    df2 = df2[[0, 1, 2]]
    df2.columns = [f'Дата {currency2_short}', f'Курс {currency2_short}', f'Время {currency2_short}']
    df2 = df2.astype({f'Курс {currency2_short}': 'float64'})
    df_dict[currency2] = df2[[f'Дата {currency2_short}', f'Курс {currency2_short}', f'Время {currency2_short}']]

    result_data = reduce(lambda x, y: x.join(y), [elem for elem in df_dict.values()])
    index_list = [i for i in range(4, len(result_data.iloc[0]), 3)]  # if currency_pairs relative RUB more then 2
    for i in index_list:
        name = result_data.columns[i - 1].split(' ')[1].split('_')[0]
        result_data[f'Результат'] = result_data.iloc[:, 1] / result_data.iloc[:, i]
    sheet_name = time.ctime().replace(':', '-')

    with pandas.ExcelWriter(file_name, engine='openpyxl', date_format='dd.mm.yyyy') as writer:
        result_data.to_excel(writer, sheet_name=sheet_name, index=False, float_format="%.4f")

    wb = openpyxl.load_workbook(file_name)
    sheet = wb[sheet_name]
    v = len(result_data)
    b = len(result_data.iloc[0])
    for i in range(1, len(result_data)+2):
        for j in range(1, len(result_data.iloc[0])):
            # width = len(sheet.cell(i, j).value) * 1.3
            sheet.column_dimensions[get_column_letter(j)].bestFit = True
            sheet.column_dimensions[get_column_letter(j)].auto_size = True
            # sheet.column_dimensions[get_column_letter(j)].width = width
            sheet.cell(i, j).alignment = Alignment(horizontal="center", vertical='center')
            if i != 1:
                sheet.cell(i, 2).number_format = '# ##0.0000" р.";-# ##0.0000" р."'
                sheet.cell(i, 5).number_format = '# ##0.0000" $";-# ##0.0000" $"'

    wb.save(file_name)

    if num_row1 in [1, 21, 31]:
        strok = "строка"
    elif num_row1 in [2, 3, 4, 22, 23, 24]:
        strok = "строки"
    else:
        strok = "строк"


    text = f'В отчете {num_row1} {strok}.'

    send_mail('Отчет MOEX', text, 'admin.vs@mail.ru', file_name)


except Exception as err:
    print(err)

finally:
    driver.close()
    exit()
